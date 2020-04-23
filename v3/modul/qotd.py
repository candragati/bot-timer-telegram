
from telegram import ParseMode
from config import *
import random
import html
from modul.kamus import kamus
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
# from openpyxl.styles import Side, Border
from datetime import datetime
import threading
from modul.setting import getUsername

lock = threading.Lock()

def qotd(update,context,acak=None):
    bot     = context.bot    
    args    = context.args    
    if acak == None and len(args) == 0:
        quote_id = None
    elif acak != None:
        quote_id = acak[0]
    elif str(args[0]).isdigit():
        quote_id = args[0]    
     
    
    chat_id = update.message["chat"]["id"]
    if quote_id == None:
        m = update.effective_message            
        try:
            waktu   = str(m.reply_to_message.date.strftime('%Y-%m-%d %H:%M:%S'))
            quote   = m.reply_to_message.text.replace("'","''")
            cek     = "SELECT nomor FROM qotd WHERE quote = '%s' AND chat_id = '%s'"%(quote,chat_id)
            barC,jumC = eksekusi(cek)
            if jumC >= 1:
                update.message.reply_text(str(kamus("qotd_dobel"))%(barC[0][0]))
            else:
                try:
                    user_name   =  update.message["reply_to_message"]["forward_from"]["username"]
                    user_id     =  update.message["reply_to_message"]["forward_from"]["id"]
                except:                        
                    user_name   =  m.reply_to_message.from_user.username
                    user_id     =  m.reply_to_message.from_user.id
                chat_type   = update.message["chat"]["type"]
                hitung      = "SELECT quote FROM qotd WHERE chat_id = '%s'"%(chat_id)
                bar, jum = eksekusi(hitung)            
                jum = jum+1
                try:
                    lock.acquire(True)
                    sql = "INSERT INTO qotd (nomor,waktu, chat_id, chat_type, user_id, user_name, quote, hapus) VALUES ('%s','%s','%s','%s','%s','%s','%s','%s')"%(
                    jum,waktu, chat_id, chat_type, user_id, user_name, quote, 0)
                    cur.execute(sql)
                    db.commit()
                finally:
                    lock.release()
                update.message.reply_text(str(kamus("quote_simpan"))%(jum,jum,jum))
        except Exception as e:            
            update.message.reply_text(str(kamus("qotd_mogok")))

    elif quote_id!=None:
        sql = "SELECT quote, user_id, nomor,user_name FROM qotd WHERE chat_id = '%s' AND nomor = '%s'"%(chat_id, quote_id)
        bar, jum = eksekusi(sql)
        if jum == 0:
            update.message.reply_text(kamus("quote_not_found"))
        else:
            # user_name = bot.get_chat(bar[0][1]).username
            # user_name = bar[0][3]
            # print (bot.get_chat(bar[0][1]))
            args        = bar[0][1],bar[0][3]
            user_name   = getUsername(update,context,args)
            teks        = "<i>{}</i>\n\n{}:{}".format(html.escape(bar[0][0]),user_name,bar[0][2])
            try:
                m_id    = update.message["reply_to_message"]["message_id"]
                bot.send_message(chat_id=update.message.chat_id,reply_to_message_id=m_id, text = teks, parse_mode=ParseMode.HTML)
            except:
                update.message.reply_html(teks)
            try:
                lock.acquire(True)
                sql_hit = "UPDATE qotd SET hit = hit+1 WHERE chat_id = '%s' AND nomor = '%s'"%(chat_id,bar[0][2])
                cur.execute(sql_hit)
                db.commit()
            finally:
                lock.release()
    else:
        update.message.reply_text("Silahkan reply atau forward chat yang akan di quote")

def rqotd(update,context):
    chat_id     = update.message["chat"]["id"]
    quote       = []
    acak        = []
    sql         = "SELECT nomor FROM qotd WHERE chat_id = '%s' AND hapus = 0"%chat_id
    bar, jum    = eksekusi(sql)
    for i in range(jum):
        quote.append(bar[i][0])
    acak.append('%s'%random.choice(quote))
    qotd(update,context,acak)

def dqotd(update,context):
    args = context.args
    chat_id = update.message["chat"]["id"]
    if len(args) == 0:
        update.message.reply_text(kamus("quote_kurang"))
    elif str(args[0]).isdigit():
        hitung = "SELECT quote FROM qotd WHERE chat_id = '%s'"%(chat_id)
        bar, jum = eksekusi(hitung)            
        if jum == 0:
            update.message.reply_text(kamus("dqotd_not_found"))
        else:
            try:
                lock.acquire(True)
                sql = ("UPDATE qotd SET hapus = 1 WHERE chat_id = '%s' AND nomor = '%s'"%(chat_id, args[0]))
                cur.execute(sql)
                db.commit()                
            finally:
                lock.release()
            update.message.reply_text(kamus("dqotd_sukses"))
    else:
        update.message.reply_text(kamus("quote_kurang"))

def xqotd(update,context):    
    bot     = context.bot
    chat_id = update.message["chat"]["id"]
    user    = update.effective_user    
    h = (
            ("ID"),
            ("WAKTU"),
            ("USERNAME"),
            ("QUOTE"),
            )
    sekarang    = datetime.now()
    hari        = datetime.strftime(sekarang.date(),"%Y-%m-%d")
    simpan      = "Export.xlsx"        
    wb          = Workbook()
    ws1         = wb.active
    ws1['A1']   = "Export QOTD"
    ws1['A1'].style = 'Title'
    # ws1['A1'].alignment = Alignment(horizontal = 'center')
    # ws1.merge_cells('A1:%s1'%chr(64+len(h)))

    ws1['A2']   = "Periode %s"%hari
    ws1['A2'].style = 'Headline 4'
    
    for col in range(len(h)):                
        ws1.cell(column = col+1, row = 3,value = h[col])
        ws1[chr(ord('A')+col)+'3'].style ='Accent4'
        ws1[chr(ord('A')+col)+'3'].alignment = Alignment(horizontal = 'center')
        ws1[chr(ord('A')+col)+'3'].font = Font(size = 8,color = 'FFFFFF')

    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['D'].width = 50
    try:    
        sql = "SELECT nomor, waktu, user_name, quote FROM qotd WHERE chat_id = '%s'"%chat_id
        bar, jum = eksekusi(sql)
        for data in range(jum):
            data0 = bar[data][0]
            data1 = bar[data][1]
            data2 = bar[data][2]
            # data2 = bot.get_chat(bar[data][2]).username
            data3 = bar[data][3]
            
            ws1.cell(column = 1,row = data+4,value = data0)
            ws1.cell(column = 2,row = data+4,value = data1)
            ws1.cell(column = 3,row = data+4,value = data2)
            ws1.cell(column = 4,row = data+4,value = data3)
        
        wb.save(filename = simpan)
        file = open("Export.xlsx","rb")
        bot.send_document(user.id, file)
            
    except Exception as e:
        update.message.reply_text(str(e))
    else:
        update.message.reply_text("Berhasil Export.\nCek PM...")

def sqotd(update,context):
    bot     = context.bot
    args    = context.args
    chat_id = update.message["chat"]["id"]
    if len(args)==0:
        hitung      = "SELECT user_id,nomor,hit,user_name FROM qotd WHERE chat_id = %s ORDER BY cast(hit as integer) DESC limit 5"%(chat_id)
        bar, jum    = eksekusi(hitung)
        teks        = "TOP Quote by hit"
        # user_name = bot.get_chat(bar[i][0]).username        
        tampil      = ''.join('x%s %s:%s\n'%(bar[i][2],bar[i][3],bar[i][1]) for i in range(jum))
    elif str(args[0]) == 'member':
        hitung      = "SELECT count(user_id),user_id,user_name FROM qotd WHERE chat_id = %s group by user_id ORDER BY count(user_id) DESC limit 5"%(chat_id)
        bar, jum    = eksekusi(hitung)        
        teks        = "Member yang quotable"
        # user_name = bot.get_chat(bar[i][0]).username        
        tampil      = ''.join('%sx %s\n'%(bar[i][0],bar[i][2]) for i in range(jum))
    else:
        user_name   = str(args[0]).replace('@','')        
        hitung      = "SELECT user_id,nomor,hit,user_name FROM qotd WHERE chat_id = ? AND user_name = ? ORDER BY cast(hit as integer) DESC limit 5"
        cur.execute(hitung,(chat_id,user_name))
        bar =  (cur.fetchall())
        jum =  (len(bar))
        if jum == 0:
            teks    = ""
            tampil  = "Silahkan mensyen member disini"
        else:
            teks    = "TOP Quote dari %s"%user_name
            # user_name = bot.get_chat(bar[i][0]).username            
            tampil  = ''.join('x%s %s:%s\n'%(bar[i][2],bar[i][3],bar[i][1]) for i in range(jum))
    update.message.reply_text('%s\n\n%s'%(teks,tampil))
